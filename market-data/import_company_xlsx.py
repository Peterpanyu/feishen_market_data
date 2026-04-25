"""
将我司电摩等 Excel 转为统一表头 CSV（与 clean_csv 输出一致），再调用 import_market_csv.py
按既有规则写入 MongoDB 集合「竞品产品」。

依赖安装（国内镜像示例）::

    pip install -r requirements.txt -i https://pypi.tuna.tsinghua.edu.cn/simple --trusted-host pypi.tuna.tsinghua.edu.cn

用法::

    py import_company_xlsx.py --input "C:\\Users\\panyu\\Desktop\\我司电摩数据汇总.xlsx"
    py import_company_xlsx.py --input "..." --dry-run
    py import_company_xlsx.py --input "..." --csv-only
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import subprocess
import sys
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from clean.schema import CANONICAL_COLUMNS  # noqa: E402

try:
    from openpyxl import load_workbook
except ImportError as e:  # pragma: no cover
    print(
        "缺少 openpyxl。请执行：pip install openpyxl "
        '-i https://pypi.tuna.tsinghua.edu.cn/simple --trusted-host pypi.tuna.tsinghua.edu.cn',
        file=sys.stderr,
    )
    raise SystemExit(1) from e


def norm_header(k: str | None) -> str:
    return re.sub(r"\s+", " ", (k or "").strip())


# 表头别名 → CANONICAL_COLUMNS 中的标准列名
_HEADER_ALIASES: dict[str, str] = {
    norm_header("产业线"): "产品线",
    norm_header("产品型号"): "型号",
    norm_header("产品名"): "型号",
    norm_header("名称"): "型号",
}


def _cell_to_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
        return ""
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        if isinstance(val, float) and val == int(val):
            return str(int(val))
        return str(val).strip()
    s = str(val).strip()
    if s.lower() == "nan":
        return ""
    return s


def _sheet_tables(path: Path, sheet_name: str | None) -> tuple[str, list[str], list[dict[str, str]]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        names = [sheet_name] if sheet_name else list(wb.sheetnames)
        best: tuple[str, list[str], list[dict[str, str]]] | None = None
        for name in names:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            raw_rows: list[list[Any]] = []
            for row in ws.iter_rows(values_only=True):
                if row and any(_cell_to_str(c) for c in row):
                    raw_rows.append(list(row))
            if len(raw_rows) < 2:
                continue
            header_cells = raw_rows[0]
            headers: list[str] = []
            for i, c in enumerate(header_cells):
                h = norm_header(_cell_to_str(c))
                headers.append(h if h else f"_列{i}")
            data_rows: list[dict[str, str]] = []
            for r in raw_rows[1:]:
                d: dict[str, str] = {}
                for j, h in enumerate(headers):
                    val = r[j] if j < len(r) else None
                    d[h] = _cell_to_str(val)
                if any(d.values()):
                    data_rows.append(d)
            if not data_rows:
                continue
            if best is None or len(data_rows) > len(best[2]):
                best = (name, headers, data_rows)
        if best is None:
            return "", [], []
        return best
    finally:
        wb.close()


def _resolve_canonical_field(excel_header: str) -> str | None:
    nh = norm_header(excel_header)
    if nh in _HEADER_ALIASES:
        return _HEADER_ALIASES[nh]
    for col in CANONICAL_COLUMNS:
        if norm_header(col) == nh:
            return col
    return None


def _row_to_canonical_row(
    row: dict[str, str],
    *,
    source_basename: str,
) -> dict[str, str]:
    out: dict[str, str] = {k: "" for k in CANONICAL_COLUMNS}
    extras: dict[str, str] = {}
    for k, v in row.items():
        if not norm_header(k):
            continue
        canon = _resolve_canonical_field(k)
        if canon and canon in CANONICAL_COLUMNS:
            if canon == "其它列_JSON":
                if v:
                    out["其它列_JSON"] = v
                continue
            if v:
                out[canon] = v
            continue
        if v:
            extras[norm_header(k)] = v
    if extras:
        raw_existing = (out.get("其它列_JSON") or "").strip()
        if raw_existing:
            try:
                parsed = json.loads(raw_existing)
                if isinstance(parsed, dict):
                    parsed.update(extras)
                    out["其它列_JSON"] = json.dumps(parsed, ensure_ascii=False, sort_keys=True)
                else:
                    out["其它列_JSON"] = json.dumps(
                        {"_表头其它列_JSON": raw_existing, **extras},
                        ensure_ascii=False,
                        sort_keys=True,
                    )
            except json.JSONDecodeError:
                out["其它列_JSON"] = json.dumps(
                    {"_表头其它列_JSON": raw_existing, **extras},
                    ensure_ascii=False,
                    sort_keys=True,
                )
        else:
            out["其它列_JSON"] = json.dumps(extras, ensure_ascii=False, sort_keys=True)
    if not (out.get("原始来源文件") or "").strip():
        out["原始来源文件"] = source_basename
    return out


def write_unified_csv(rows: list[dict[str, str]], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(CANONICAL_COLUMNS), extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in CANONICAL_COLUMNS})


def main() -> int:
    p = argparse.ArgumentParser(description="Excel → 统一 CSV → import_market_csv（竞品产品）")
    p.add_argument("--input", type=Path, required=True, help="xlsx 路径")
    p.add_argument(
        "--output",
        type=Path,
        default=None,
        help="统一 CSV 输出路径（默认 data/cleaned/<xlsx 文件名>_统一.csv）",
    )
    p.add_argument("--sheet", default=None, help="仅读取指定工作表名；不传则自动选数据行最多的表")
    p.add_argument("--csv-only", action="store_true", help="只写 CSV，不执行 import_market_csv.py")
    p.add_argument("--dry-run", action="store_true", help="传给 import_market_csv.py")
    p.add_argument("--wipe", action="store_true", help="传给 import_market_csv.py")
    p.add_argument(
        "--delete-product-line",
        default=None,
        metavar="产品线",
        help="传给 import_market_csv.py",
    )
    p.add_argument("--encoding", default=None, help="传给 import_market_csv.py（CSV 编码，一般不必）")
    args = p.parse_args()

    inp: Path = args.input.expanduser()
    if not inp.is_file():
        print(f"文件不存在：{inp}", file=sys.stderr)
        return 2

    used_sheet, _headers, data_rows = _sheet_tables(inp, args.sheet)
    if not data_rows:
        print(
            f"未在「{inp.name}」中读到可用数据：请确认至少一张工作表有「表头行 + 至少一行数据」，"
            "且单元格非空。当前文件各表可能为空或仅一行。",
            file=sys.stderr,
        )
        return 2

    base = inp.name
    out_csv = args.output or (ROOT / "data" / "cleaned" / f"{inp.stem}_统一.csv")
    canon_rows = [_row_to_canonical_row(r, source_basename=base) for r in data_rows]
    write_unified_csv(canon_rows, out_csv)
    print(f"已使用工作表：{used_sheet!r}，数据行：{len(canon_rows)}，统一 CSV：{out_csv}")

    if args.csv_only:
        print("已跳过 Mongo 导入（--csv-only）。需要时执行：")
        print(f'  py import_market_csv.py --csv "{out_csv}"')
        return 0

    cmd = [sys.executable, str(ROOT / "import_market_csv.py"), "--csv", str(out_csv)]
    if args.dry_run:
        cmd.append("--dry-run")
    if args.wipe:
        cmd.append("--wipe")
    if args.delete_product_line:
        cmd.extend(["--delete-product-line", args.delete_product_line])
    if args.encoding:
        cmd.extend(["--encoding", args.encoding])

    r = subprocess.run(cmd, cwd=str(ROOT))
    return int(r.returncode)


if __name__ == "__main__":
    raise SystemExit(main())
